#
# Convert Microsoft Project .mpp file to Excel Gantt Chart
#

import datetime
import json
import os

import openpyxl
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.utils.cell import get_column_letter

import tab_complete

# global variables
path_to_jvm = None  # path to Java VM (jvm.dll)
fontname    = None  # font name
tab_title   = None  # tab title
start_gantt = None  # start date of gantt chart
end_gantt   = None  # end date of gantt chart
holidays    = None  # list of holidays

def is_holiday(date):
    """
    Check if the given date is a holiday or weekend.

    Args:
        date (datetime.date): The date to check.
    """

    w = date.weekday()  # day of week (0:monday - 6:sunday)
    if w == 5 or w == 6:
        return True
    
    for h in holidays:
        if date == h:
            return True

    return False

def set_title_row(ws):
    """
    Set title row and column width for gantt chart template.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
    """

    # set height of row
    # ws.row_dimensions[1].height = 40  # Title row

    # set column width
    ws.column_dimensions['A'].width =  8  # Task #
    ws.column_dimensions['B'].width = 50  # Subject
    ws.column_dimensions['C'].width = 16  # Assigned
    ws.column_dimensions['D'].width = 12  # Start Date
    ws.column_dimensions['E'].width = 12  # Due Date
    ws.column_dimensions['F'].width = 12  # Closed Date
    ws.column_dimensions['G'].width = 12  # Done Ratio

    ws.cell(1, 1).value = '#'
    ws.cell(1, 1).font = Font(name=fontname)
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 2).value = 'Subject'
    ws.cell(1, 2).font = Font(name=fontname)
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 3).value = 'Assigned'
    ws.cell(1, 3).font = Font(name=fontname)
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 4).value = 'Start'
    ws.cell(1, 4).font = Font(name=fontname)
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 5).value = 'Due'
    ws.cell(1, 5).font = Font(name=fontname)
    ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 6).value = 'Closed'
    ws.cell(1, 6).font = Font(name=fontname)
    ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 7).value = 'Done(%)'
    ws.cell(1, 7).font = Font(name=fontname)
    ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')

    # merge cells for title row
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')

def excel_set_GanttChart_date(ws):
    """
    Set month and day for gantt chart in excel.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
    """

    # fill color for holidays
    fillLightPink = PatternFill(patternType='solid', fgColor='ffccff')  # Light Pink

    column = 8  # H -
    d = start_gantt
    while d <= end_gantt:
        ws.column_dimensions[ get_column_letter(column) ].width = 4

        # Month
        if d == start_gantt or d.day == 1:
            ws.cell(1, column, d)
            ws.cell(1, column).number_format = 'mm'
            ws.cell(1, column).font = Font(name=fontname)
            ws.cell(1, column).alignment = Alignment(horizontal='center', vertical='center')

        # Day
        ws.cell(2, column, d)
        ws.cell(2, column).number_format = 'dd'
        ws.cell(2, column).font = Font(name=fontname)
        ws.cell(2, column).alignment = Alignment(horizontal='center', vertical='center')

        # fill on holiday column
        if is_holiday(ws.cell(2, column).value):
            ws.cell(2, column).fill = fillLightPink

        d += datetime.timedelta(days=1)
        column += 1

def set_task_format(ws, row):
    """
    Set format for each task row in gantt chart.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        row (int): row number for the task
    """

    ws.cell(row, 1).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 1).font = Font(name=fontname)
    ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 2).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 2).font = Font(name=fontname)

    ws.cell(row, 3).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 3).font = Font(name=fontname)
    ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 4).number_format = 'yyyy/mm/dd'
    ws.cell(row, 4).font = Font(name=fontname)
    ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 5).number_format = 'yyyy/mm/dd'
    ws.cell(row, 5).font = Font(name=fontname)
    ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 6).number_format = 'yyyy/mm/dd'
    ws.cell(row, 6).font = Font(name=fontname)
    ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 7).number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
    ws.cell(row, 7).font = Font(name=fontname)
    ws.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')

def set_conditional_format(ws, min_row, max_row):
    """
    Set conditional formatting for gantt chart template.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        min_row (int): minimum row number for gantt chart
        max_row (int): maximum row number for gantt chart
    """

    # progress bar : F
    r1 = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='31869B', showValue=True, minLength=0, maxLength=100)
    range = f'$G${min_row}:$G${max_row}'
    ws.conditional_formatting.add(range, r1)

    # gantt chart : H -
    # count the number of date columns and find the last column
    start_gantt_column = 8  # H -
    days = nColumnGantt = end_gantt - start_gantt
    nColumnGantt = days.days
    end_gantt_column = start_gantt_column + nColumnGantt

    # condition 1 : completed part considering progress percentage
    c1 = '=AND( $D3<=H$2, H$2<=ROUNDDOWN( ($E3-$D3+1)*$G3, 0 )+$D3-1 )'
    # condition 2 : uncompleted part considering progress percentage
    c2 = '=AND( $D3<=H$2, H$2<=$E3 )'
    # condition 3 : task for future
    c3 = '=AND( $D3<=H$2, H$2<=$E3, TODAY()<H$2 )'
    # condition 4 : today
    c4 = '=AND( H$2=TODAY() )'
    # condition 5 : overdue (due cells)
    c5 = '=AND( $E3<>"", $E3<TODAY(), $G3<1 )'

    # fromat 1 : fill completed part
    f1 = PatternFill(patternType='solid', bgColor='8888ff')
    # formay 2 : fill uncompleted part
    f2 = PatternFill(patternType='solid', bgColor='ff8888')
    # format 3 : future task
    f3 = PatternFill(patternType='solid', bgColor='cccccc')
    # format 4 : today
    f4 = PatternFill(patternType='lightGray', fgColor='31869b')
    # format 5 : overdue (due cells)
    f5 = PatternFill(patternType='solid', bgColor='ffff88')

    # combine conditions and formats
    r1 = FormulaRule(formula=[c1] , stopIfTrue=None, fill=f1)
    r2 = FormulaRule(formula=[c2] , stopIfTrue=None, fill=f2)
    r3 = FormulaRule(formula=[c3] , stopIfTrue=None, fill=f3)
    r4 = FormulaRule(formula=[c4] , stopIfTrue=None, fill=f4)
    r5 = FormulaRule(formula=[c5] , stopIfTrue=None, fill=f5)

    # set conditional format
    start_cell = f'${'H'}${min_row}'
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r1)
    ws.conditional_formatting.add(cells, r2)
    ws.conditional_formatting.add(cells, r3)
    start_cell = f'${'H'}${min_row-1}' # (-1) because including month row
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r4)
    start_cell = f'${'E'}${min_row}' # from due date column
    end_cell   = f'${'E'}${max_row}' # to due date column
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r5)

    # fill holiday cells
    r = min_row
    fillLightPink = PatternFill(patternType='solid', fgColor='ffdcff')  # Light Pink
    side = Side(style='thin', color='aaaaaa')
    border = Border(top=side, bottom=side, left=side, right=side)
    while r <= max_row:
        # set_task_format(ws, r)
        c = start_gantt_column
        while c <= end_gantt_column:
            v = ws.cell(2, c).value
            if is_holiday(v):
                ws.cell(r, c).fill = fillLightPink
            ws.cell(r, c).border = border  # set border line to all cells in gantt chart area
            c += 1
        r += 1

def write_task(ws, row, indent, taskid, taskname, resourcenames, start, finish, patient=False):
    """
    Write task information to the worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        row (int): row number for the task
        indent (int): nested level
        taskid (str): task ID
        taskname (str): task name
        resourcenames (list): list of assigned resource names
        start (str): start date of the task
        finish (str): finish date of the task
        patient (bool): it is a patient task or not (default: False)
    """

    # replace - to / for date format, and convert string to datetime.date
    if start:
        start = start.replace('-', '/')
        start = datetime.datetime.strptime(start, '%Y/%m/%d').date()
    if finish:
        finish = finish.replace('-', '/')
        finish = datetime.datetime.strptime(finish, '%Y/%m/%d').date()

    ws.cell(row, 1).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 1).value = taskid
    ws.cell(row, 1).font = Font(name=fontname)
    ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 2).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 2).value = taskname
    ws.cell(row, 2).font = Font(name=fontname, bold=True if patient else False)
    ws.cell(row, 2).alignment = Alignment(indent=indent, vertical='center')

    ws.cell(row, 3).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 3).value = resourcenames
    ws.cell(row, 3).font = Font(name=fontname)
    ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 4).number_format = 'yyyy/mm/dd'
    ws.cell(row, 4).value = start
    ws.cell(row, 4).font = Font(name=fontname)
    ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 5).number_format = 'yyyy/mm/dd'
    ws.cell(row, 5).value = finish
    ws.cell(row, 5).font = Font(name=fontname)
    ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 6).number_format = 'yyyy/mm/dd'
    ws.cell(row, 6).value = '' # No 'closed_on' information in mpp file
    ws.cell(row, 6).font = Font(name=fontname)
    ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 7).number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
    ws.cell(row, 7).value = 0 # No 'done_ratio' information in mpp file
    ws.cell(row, 7).font = Font(name=fontname)
    ws.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')

    # Regardless of the value of done_ratio, if closed_on exists, set to 100%
    # if hasattr(issue, 'closed_on') and issue.closed_on is not None:
    #     ws.cell(row, 7).value = 1.0  # 100% complete

def enum_tasks(mpp, ws, row):
    """
    Enumerate tasks from a Microsoft Project .mpp file and write them to the worksheet.
    This function uses the JPype library to read the .mpp file and extract task information.

    Args:
        mpp (str): Microsoft Project .mpp file name
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        row (int): starting row number for writing tasks

    Returns:
        int: The next row number after writing all tasks.
    """

    import os

    jvm = path_to_jvm
    # Check if the "JAVA_HOME" environment variable is NOT set.
    if os.environ.get('JAVA_HOME') is None:
        # If the "path_to_jvm" is set and it is a valid.
        if path_to_jvm is None or not os.path.exists(path_to_jvm):
            # Unable to continue.
            print("Error: 'path_to_jvm' entry in the config.json is not set or invalid.")
            print("Please set 'path_to_jvm' entry in the config.json or as the 'JAVA_HOME' environment variable to the path of jvm.dll.")
            exit(1)
    else:
        # "JAVA_HOME" is set, so use it.
        jvm = None

    import jpype
    import mpxj

    try:
        jpype.startJVM(jvmpath=jvm)
        from org.mpxj.reader import UniversalProjectReader
    except jpype.JVMNotFoundException:
        print("Error: Unable to load Java VM. Please set 'path_to_jvm' entry in the config.json or as the 'JAVA_HOME' environment variable to the path of jvm.dll.")
        exit(1)

    project = UniversalProjectReader().read(mpp)

    tasks = project.getTasks()
    # How many tasks are there?
    print(f'Total number of tasks: {len(tasks)}')

    def print_task_tree(task, row, level=0):
        # Do not display level=0 (root task), start from child tasks
        if level > 0:
            indent = level-1
            id     = str(task.getID())
            name   = str(task.getName())
            child_tasks = list(task.getChildTasks())
            # Get assigned resource names as a list
            assignments = task.getResourceAssignments()
            resource_names = []
            for assignment in assignments:
                resource = assignment.getResource()
                if resource is not None:
                    resource_name = str(resource.getName())
                    if resource_name is not None:
                        resource_names.append(resource_name)
            # Display resource names in parentheses
            resource_str = f" {', '.join(resource_names)}" if resource_names else ''
            if child_tasks:
                # If there are child tasks, do not display dates
                write_task(ws, row, indent, id, name, resource_str, start='', finish='', patient=True)
            else:
                start  = str(task.getStart())
                finish = str(task.getFinish())
                write_task(ws, row, indent, id, name, resource_str, start[:10], finish[:10], patient=False)
            row += 1

        # Recursively display child tasks
        for child in task.getChildTasks():
            row = print_task_tree(child, row, level + 1)

        return row

    # Extract only root tasks and display them recursively
    for task in tasks:
        if task.getParentTask() is None:
            row = print_task_tree(task, row)

    jpype.shutdownJVM()

    return row

def main(mpp):
    """
    Main function to generate gantt chart template in excel.

    Args:
        mpp (str): Microsoft Project .mpp file name
    """

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # tab title
    ws.title = tab_title

    # title row
    set_title_row(ws)

    # set month and day for gantt chart
    excel_set_GanttChart_date(ws)

    start_row = 3
    end_row = enum_tasks(mpp, ws, start_row) - 1

    # freeze panes
    ws.freeze_panes = 'H3'
    # set filter
    ws.auto_filter.ref = f'A2:G{end_row}'

    # conditional formatting
    set_conditional_format(ws, start_row, end_row)

    while True:
        print(f"Input file name (It doesn't need '.xlsx' extention.) : ", end='')
        f = input()
        # check if file name is empty
        if f == '':
            print("File name can't be empty.")
            continue
        # confirm oberwrite if file exists
        if os.path.exists(f'.\\{f}.xlsx'):
            print(f"'{f}.xlsx' already exists. Do you want to overwrite it? [y/_] : ", end='')
            yn = input().upper()
            if yn != 'Y':
                continue
        try:
            wb.save(f'.\\{f}.xlsx')
            break
        except:
            print(f"Error : Can't save to '{f}.xlsx")
            print(f"Do you want to try again? [_/n] : ", end='')
            yn = input().upper()
            if yn == 'N':
                break 

def load_config_from_json():
    """
    Load configuration from 'config.json'.
    """

    config_file = 'config.json' # constant file name
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            try:
                config_path_to_jvm = config.get('path_to_jvm', None)
                config_font_name   = config['font_name']
                config_tab_title   = config['tab_title']
                config_start_date  = config['start_date']
                config_end_date    = config['end_date']
                config_holidais    = config['holidays']
            except KeyError as e:
                print(f'format error in config.json: {e}')
                return False
    else:
        print(f"config file '{config_file}' not found.")
        return False

    try:
        global path_to_jvm, fontname, tab_title, start_gantt, end_gantt, holidays
        path_to_jvm  = config_path_to_jvm
        fontname     = config_font_name
        tab_title    = config_tab_title
        start_gantt  = datetime.datetime.strptime(config_start_date, '%Y/%m/%d').date()
        end_gantt    = datetime.datetime.strptime(config_end_date, '%Y/%m/%d').date()
        holidays     = [datetime.datetime.strptime(date, '%Y/%m/%d').date() for date in config_holidais]
    except ValueError as e:
        print(f'format error in config.json: {e}')
        return False

    return True

if __name__ == '__main__':
    if load_config_from_json():
        print(f"Input Microsoft project .mpp file name : ", end='')
        mpp = input()
        if not os.path.exists(mpp):
            print(f"File '{mpp}' not found.")
        else:
            main(mpp)

